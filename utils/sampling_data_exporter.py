"""
Quadrat Data Processing and Excel Export Module

This module is used to process quadrat data in remote sensing images and export the results to Excel files.

Function Description:
1. Quadrat Area Determination:
   - After determining the center coordinate point of the quadrat, use ArcMap's "Buffer" tool to generate a square quadrat area of 1m*1m 
     (or other sizes) near the center point (quadrats can be distinguished by numbering)

2. Index Statistics:
   - After determining the quadrat area, use ArcMap's "Zonal Statistics" tool to extract reflectance and calculate spectral indices for the quadrat area
   - Supported spectral indices:
     * NDVI = (NIR - Red) / (NIR + Red)
     * GNDVI = (NIR - Green) / (NIR + Green)
     * NDRE = (NIR - RedEdge) / (NIR + RedEdge)

Exported data includes:
1. Quadrat coordinate points (four corner coordinates + center point coordinate)
2. Quadrat reflectance: maximum, minimum, average, and standard deviation of each band (Red, Green, Blue, RE, NIR) within the quadrat
3. Quadrat spectral indices: maximum, minimum, average, and standard deviation of NDVI, GNDVI, and NDRE
"""

import io
import os
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

# Define Excel column headers
COLUMNS = [
    '样方名称', '坐标点', '坐标中心点',
    'NDVI最大值', 'NDVI最小值', 'NDVI平均值', 'NDVI标准差',
    'GNDVI最大值', 'GNDVI最小值', 'GNDVI平均值', 'GNDVI标准差',
    'NDRE最大值', 'NDRE最小值', 'NDRE平均值', 'NDRE标准差',
    'Red最大值', 'Red最小值', 'Red平均值', 'Red标准差',
    'Green最大值', 'Green最小值', 'Green平均值', 'Green标准差',
    'RE最大值', 'RE最小值', 'RE平均值', 'RE标准差',
    'NIR最大值', 'NIR最小值', 'NIR平均值', 'NIR标准差'
]

# Define the starting column index for each algorithm in Excel
COLUMN_CONFIG = {
    'ndvi': 5,
    'gndvi': 9,
    'ndre': 13,
    'r': 17,
    'g': 21,
    're': 25,
    'n': 29
}
TEMPLATE_XLSX_PATH = Path(os.getcwd()) / "docker/rsdm/template.xlsx"


def format_coord(coords) -> str:
    """
    Format coordinate point as string with 6 decimal places
    """
    def process(coord):
        if isinstance(coord, list) and len(coord) == 2:
            return "{:.6f}, {:.6f}".format(coord[0], coord[1])
        return str(coord)

    coord_format = [process(c) for c in coords]

    return '\n'.join(coord_format)


def create_template():
    """
    Create Excel template file
    """
    df = pd.DataFrame(columns=COLUMNS)
    df.to_excel(TEMPLATE_XLSX_PATH, index=False)


def generate_excel(quadrat_records) -> io.BytesIO:
    """
    Generate Excel file based on quadrat records
    
    Args:
        quadrat_records: List of quadrat records
        
    Returns:
        io.BytesIO: Byte stream containing Excel data
    """
    # Read template file
    wb = load_workbook(TEMPLATE_XLSX_PATH)
    ws = wb.active

    # Process each quadrat record
    for row_idx, quadrat in enumerate(quadrat_records, 2):
        # Write basic quadrat information
        ws.cell(row=row_idx, column=1, value=quadrat.sort_no if quadrat.sort_no else '样方{}'.format(row_idx - 1))
        # ws.cell(row=row_idx, column=1, value=quadrat.sort_no)
        ws.cell(row=row_idx, column=2, value=quadrat.name)
        ws.cell(row=row_idx, column=3, value=format_coord(quadrat.coords))
        ws.cell(row=row_idx, column=4, value=format_coord(quadrat.center))

        # Build statistics dictionary
        stats_dict = {stat.algo_name.lower(): stat for stat in quadrat.statistics}

        # Write algorithm statistics
        for algo, start_col in COLUMN_CONFIG.items():
            stat = stats_dict.get(algo, None)

            if stat:
                max_cell = ws.cell(row=row_idx, column=start_col, value=stat.dn_max)
                min_cell = ws.cell(row=row_idx, column=start_col + 1, value=stat.dn_min)
                mean_cell = ws.cell(row=row_idx, column=start_col + 2, value=stat.dn_mean)
                std_cell = ws.cell(row=row_idx, column=start_col + 3, value=stat.dn_std)

                # Set number format to 10 decimal places
                max_cell.number_format = '0.0000000000'
                min_cell.number_format = '0.0000000000'
                mean_cell.number_format = '0.0000000000'
                std_cell.number_format = '0.0000000000'

    # Save to byte stream and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output
